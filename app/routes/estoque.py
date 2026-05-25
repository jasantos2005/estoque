import sqlite3, csv, io
from pathlib import Path
from datetime import datetime
from fastapi import APIRouter, Header, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional
from jose import jwt, JWTError

DB_PATH   = Path(__file__).parent.parent.parent / "data" / "estoque.db"
SECRET    = "hub_estoque_secret_trocar_depois"
ALGORITMO = "HS256"

router = APIRouter()

def get_db():
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn

def verificar_token(authorization: str = ""):
    if not authorization.startswith("Bearer "):
        raise HTTPException(401, "Token inválido")
    try:
        return jwt.decode(authorization[7:], SECRET, algorithms=[ALGORITMO])
    except JWTError:
        raise HTTPException(401, "Token expirado")

def calcular_dias(saldo, consumo_dia):
    if not consumo_dia or consumo_dia <= 0:
        return 999
    return int(saldo / consumo_dia)

def get_itens(categoria_like: str, de: str, ate: str):
    conn = get_db()
    cur  = conn.cursor()
    cur.execute("""
        SELECT p.id_produto, p.descricao, p.categoria, p.unidade,
               COALESCE(p.estoque_minimo, 0) as estoque_minimo,
               COALESCE(s.saldo, 0) as saldo,
               COALESCE(m.saida_periodo, 0) as saida_periodo
        FROM produtos p
        LEFT JOIN saldos s ON s.id_produto = p.id_produto
        LEFT JOIN (
            SELECT id_produto, SUM(quantidade) as saida_periodo
            FROM movimentacoes
            WHERE tipo='saida'
              AND (? = '' OR data >= ?)
              AND (? = '' OR data <= ?)
            GROUP BY id_produto
        ) m ON m.id_produto = p.id_produto
        WHERE p.categoria LIKE ?
        ORDER BY p.descricao
    """, (de, de, ate, ate, categoria_like))
    rows = cur.fetchall()
    conn.close()
    result = []
    for r in rows:
        if float(r["saldo"]) <= 0:
            continue
        consumo_dia = r["saida_periodo"] / 90 if r["saida_periodo"] else 0
        dias = calcular_dias(r["saldo"], consumo_dia)
        result.append({
            "id_produto":      r["id_produto"],
            "descricao":       r["descricao"],
            "categoria":       r["categoria"],
            "unidade":         r["unidade"],
            "estoque_minimo":  round(float(r["estoque_minimo"]), 2),
            "saldo":           round(float(r["saldo"]), 2),
            "saida_periodo":   round(float(r["saida_periodo"]), 2),
            "consumo_dia":     round(consumo_dia, 2),
            "dias_cobertura":  dias,
        })
    return result

@router.get("/dashboard")
def dashboard(de: str = "", ate: str = "", authorization: str = Header("")):
    verificar_token(authorization)
    conn = get_db()
    cur  = conn.cursor()
    cur.execute("""
        SELECT p.id_produto, p.descricao, p.categoria, p.unidade,
               COALESCE(s.saldo, 0) as saldo,
               COALESCE(m.saida_periodo, 0) as saida_periodo
        FROM produtos p
        LEFT JOIN saldos s ON s.id_produto = p.id_produto
        LEFT JOIN (
            SELECT id_produto, SUM(quantidade) as saida_periodo
            FROM movimentacoes
            WHERE tipo='saida'
              AND (? = '' OR data >= ?)
              AND (? = '' OR data <= ?)
            GROUP BY id_produto
        ) m ON m.id_produto = p.id_produto
    """, (de, de, ate, ate))
    rows = cur.fetchall()

    # pedidos pendentes
    cur.execute("SELECT COUNT(*) as c FROM pedidos_compra WHERE status='pendente'")
    ped_row = cur.fetchone()
    pedidos_pendentes = ped_row["c"] if ped_row else 0
    conn.close()

    itens_criticos = []
    top_consumo = []
    dias_casa = []; dias_infra = []
    total_casa = total_infra = 0

    itens_zerados = []
    itens_alerta  = []
    itens_parados = []
    total_saida   = 0.0
    dist_casa     = {"critico": 0, "alerta": 0, "ok": 0, "parado": 0, "zerado": 0}
    dist_infra    = {"critico": 0, "alerta": 0, "ok": 0, "parado": 0, "zerado": 0}

    for r in rows:
        saldo      = float(r["saldo"])
        saida      = float(r["saida_periodo"])
        consumo_dia = saida / 90 if saida else 0
        dias        = calcular_dias(saldo, consumo_dia)
        cat         = (r["categoria"] or "GERAL").upper()
        dist        = dist_casa if cat == "CASA" else dist_infra if cat == "INFRA" else None

        total_saida += saida

        if cat == "CASA":
            total_casa += 1
            if dias < 999: dias_casa.append(dias)
        elif cat == "INFRA":
            total_infra += 1
            if dias < 999: dias_infra.append(dias)

        base = {
            "id_produto":     r["id_produto"],
            "descricao":      r["descricao"],
            "categoria":      cat,
            "unidade":        r["unidade"],
            "saldo":          round(saldo, 2),
            "saida_periodo":  round(saida, 2),
            "consumo_dia":    round(consumo_dia, 2),
            "dias_cobertura": dias,
        }

        if saldo <= 0:
            itens_zerados.append(base)
            if dist: dist["zerado"] += 1
            continue

        if dias < 10:
            itens_criticos.append(base)
            if dist: dist["critico"] += 1
        elif dias < 20:
            itens_alerta.append(base)
            if dist: dist["alerta"] += 1
        elif saida == 0:
            itens_parados.append(base)
            if dist: dist["parado"] += 1
        else:
            if dist: dist["ok"] += 1

        if saida > 0:
            top_consumo.append({
                "id_produto":    r["id_produto"],
                "descricao":     r["descricao"],
                "categoria":     cat,
                "unidade":       r["unidade"],
                "saida_periodo": round(saida, 2),
                "consumo_dia":   round(consumo_dia, 2),
                "dias_cobertura": dias,
            })

    top_consumo    = sorted(top_consumo, key=lambda x: x["saida_periodo"], reverse=True)[:10]
    itens_criticos = sorted(itens_criticos, key=lambda x: x["dias_cobertura"])
    itens_alerta   = sorted(itens_alerta,   key=lambda x: x["dias_cobertura"])

    cob_casa  = int(sum(dias_casa)  / len(dias_casa))  if dias_casa  else 0
    cob_infra = int(sum(dias_infra) / len(dias_infra)) if dias_infra else 0
    rup_casa  = round(len([d for d in dias_casa  if d < 5]) / max(len(dias_casa), 1)  * 100)
    rup_infra = round(len([d for d in dias_infra if d < 5]) / max(len(dias_infra), 1) * 100)
    consumo_dia_total = round(total_saida / 90, 2)

    return {
        "resumo": {
            "cobertura_casa":    cob_casa,
            "cobertura_infra":   cob_infra,
            "ruptura_pct_casa":  rup_casa,
            "ruptura_pct_infra": rup_infra,
            "itens_criticos":    itens_criticos,
            "itens_alerta":      itens_alerta,
            "itens_zerados":     itens_zerados,
            "itens_parados":     len(itens_parados),
            "pedidos_pendentes": pedidos_pendentes,
            "total_produtos":    len(rows),
            "total_saida_periodo": round(total_saida, 2),
            "consumo_dia_total": consumo_dia_total,
            "total_casa":        total_casa,
            "total_infra":       total_infra,
            "top_consumo":       top_consumo,
            "dist_casa":         dist_casa,
            "dist_infra":        dist_infra,
        }
    }

@router.get("/casa")
def estoque_casa(de: str = "", ate: str = "", authorization: str = Header("")):
    verificar_token(authorization)
    return {"itens": get_itens("%CASA%", de, ate)}

@router.get("/infra")
def estoque_infra(de: str = "", ate: str = "", authorization: str = Header("")):
    verificar_token(authorization)
    return {"itens": get_itens("%INFRA%", de, ate)}

@router.get("/sugestao")
def sugestao(de: str = "", ate: str = "", authorization: str = Header("")):
    verificar_token(authorization)
    conn = get_db()
    cur  = conn.cursor()
    # Só produtos que tiveram movimentação real (saida pelo sync_ixc ou manual)
    cur.execute("""
        SELECT p.id_produto, p.descricao, p.categoria, p.unidade,
               COALESCE(s.saldo, 0) as saldo,
               COALESCE(m.saida_periodo, 0) as saida_periodo
        FROM produtos p
        LEFT JOIN saldos s ON s.id_produto = p.id_produto
        INNER JOIN (
            SELECT id_produto, SUM(quantidade) as saida_periodo
            FROM movimentacoes
            WHERE tipo='saida'
            GROUP BY id_produto
        ) m ON m.id_produto = p.id_produto
        WHERE m.saida_periodo > 0
    """)
    rows = cur.fetchall()
    conn.close()
    result = []
    for r in rows:
        if float(r["saldo"]) < 0:
            continue
        consumo_dia = float(r["saida_periodo"]) / 90 if r["saida_periodo"] else 0
        dias = calcular_dias(float(r["saldo"]), consumo_dia)
        if dias < 20:
            qtd_sugerida = max(20, int(consumo_dia * 30 * 2 - float(r["saldo"])))
            result.append({
                "id_produto":     r["id_produto"],
                "descricao":      r["descricao"],
                "categoria":      r["categoria"],
                "unidade":        r["unidade"],
                "saldo":          round(float(r["saldo"]), 2),
                "consumo_dia":    round(consumo_dia, 2),
                "dias_cobertura": dias,
                "qtd_sugerida":   qtd_sugerida,
            })
    return {"itens": sorted(result, key=lambda x: x["dias_cobertura"])}

@router.get("/movimentacoes")
def listar_movimentacoes(de: str = "", ate: str = "", authorization: str = Header("")):
    verificar_token(authorization)
    conn = get_db()
    cur  = conn.cursor()
    cur.execute("""
        SELECT m.id, m.id_produto, p.descricao, m.tipo, m.quantidade,
               m.responsavel, m.obs, m.data
        FROM movimentacoes m
        LEFT JOIN produtos p ON p.id_produto = m.id_produto
        WHERE (? = '' OR m.data >= ?)
          AND (? = '' OR m.data <= ?)
        ORDER BY m.id DESC
        LIMIT 200
    """, (de, de, ate, ate))
    rows = cur.fetchall()
    conn.close()
    return {"movimentacoes": [dict(r) for r in rows]}

class MovimentacaoBody(BaseModel):
    tipo: str
    id_produto: str
    quantidade: float
    obs: Optional[str] = ""
    responsavel: Optional[str] = "manual"

@router.post("/movimentacao")
def criar_movimentacao(body: MovimentacaoBody, authorization: str = Header("")):
    payload = verificar_token(authorization)
    conn = get_db()
    cur  = conn.cursor()
    prod = cur.execute("SELECT id_produto FROM produtos WHERE id_produto=?", (body.id_produto,)).fetchone()
    if not prod:
        conn.close()
        raise HTTPException(404, f"Produto '{body.id_produto}' não encontrado")
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cur.execute("""
        INSERT INTO movimentacoes (id_produto, tipo, quantidade, responsavel, obs, data)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (body.id_produto, body.tipo, body.quantidade,
          body.responsavel or payload.get("username", "manual"), body.obs or "", agora))
    # Atualiza saldo
    if body.tipo == "entrada":
        cur.execute("INSERT INTO saldos (id_produto, saldo) VALUES (?,?) ON CONFLICT(id_produto) DO UPDATE SET saldo=saldo+?",
                    (body.id_produto, body.quantidade, body.quantidade))
    elif body.tipo == "saida":
        cur.execute("UPDATE saldos SET saldo = MAX(0, saldo - ?) WHERE id_produto=?",
                    (body.quantidade, body.id_produto))
    conn.commit()
    conn.close()
    return {"ok": True, "msg": "Movimentação registrada"}

@router.get("/casa/csv")
def csv_casa(de: str = "", ate: str = "", authorization: str = Header("")):
    verificar_token(authorization)
    return _gerar_csv(get_itens("%CASA%", de, ate), "estoque_casa")

@router.get("/infra/csv")
def csv_infra(de: str = "", ate: str = "", authorization: str = Header("")):
    verificar_token(authorization)
    return _gerar_csv(get_itens("%INFRA%", de, ate), "estoque_infra")

@router.get("/dashboard/csv")
def csv_dashboard(de: str = "", ate: str = "", authorization: str = Header("")):
    verificar_token(authorization)
    itens = get_itens("%", de, ate)
    return _gerar_csv(itens, "estoque_completo")

def _gerar_csv(itens, nome):
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=["id_produto","descricao","categoria","unidade","saldo","saida_periodo","consumo_dia","dias_cobertura"])
    w.writeheader()
    w.writerows(itens)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={nome}.csv"}
    )


@router.get("/produtos-consumo")
def produtos_consumo(authorization: str = Header("")):
    verificar_token(authorization)
    conn = get_db()
    cur  = conn.cursor()
    cur.execute("""
        SELECT p.id_produto, p.descricao, p.categoria, p.unidade,
               COALESCE(s.saldo, 0) as saldo,
               COALESCE(m.saida_total, 0) as saida_total
        FROM produtos p
        LEFT JOIN saldos s ON s.id_produto = p.id_produto
        INNER JOIN (
            SELECT id_produto, SUM(quantidade) as saida_total
            FROM movimentacoes WHERE tipo='saida'
            GROUP BY id_produto
        ) m ON m.id_produto = p.id_produto
        WHERE m.saida_total > 0 AND COALESCE(s.saldo, 0) >= 0
        ORDER BY p.descricao
    """)
    rows = cur.fetchall()
    conn.close()
    result = []
    for r in rows:
        consumo_dia = float(r["saida_total"]) / 90
        dias = calcular_dias(float(r["saldo"]), consumo_dia)
        result.append({
            "id_produto":     r["id_produto"],
            "descricao":      r["descricao"],
            "categoria":      r["categoria"] or "GERAL",
            "unidade":        r["unidade"] or "un",
            "saldo":          round(float(r["saldo"]), 2),
            "saida_total":    round(float(r["saida_total"]), 2),
            "consumo_dia":    round(consumo_dia, 2),
            "dias_cobertura": dias,
        })
    return {"itens": result}


class EstoqueMinimoBody(BaseModel):
    id_produto: str
    estoque_minimo: float

@router.post("/estoque-minimo")
def set_estoque_minimo(body: EstoqueMinimoBody, authorization: str = Header("")):
    verificar_token(authorization)
    conn = get_db()
    cur  = conn.cursor()
    prod = cur.execute("SELECT id_produto FROM produtos WHERE id_produto=?", (body.id_produto,)).fetchone()
    if not prod:
        conn.close()
        raise HTTPException(404, "Produto nao encontrado")
    cur.execute("UPDATE produtos SET estoque_minimo=? WHERE id_produto=?",
                (body.estoque_minimo, body.id_produto))
    conn.commit()
    conn.close()
    return {"ok": True}


@router.post("/recalcular-minimo")
def recalcular_minimo(authorization: str = Header("")):
    verificar_token(authorization)
    conn = get_db()
    cur  = conn.cursor()
    rows = cur.execute("""
        SELECT id_produto, SUM(quantidade)/90.0 as consumo_dia
        FROM movimentacoes WHERE tipo='saida'
        GROUP BY id_produto HAVING consumo_dia > 0
    """).fetchall()
    atualizados = 0
    for r in rows:
        est_min = round(float(r["consumo_dia"]) * 20, 2)
        cur.execute(
            "UPDATE produtos SET estoque_minimo=? WHERE id_produto=?",
            (est_min, r["id_produto"])
        )
        atualizados += 1
    conn.commit()
    conn.close()
    return {"ok": True, "atualizados": atualizados, "msg": f"{atualizados} produtos atualizados"}


@router.get("/historico-produto/{id_produto}")
def historico_produto(id_produto: str, authorization: str = Header("")):
    verificar_token(authorization)
    conn = get_db()
    cur  = conn.cursor()
    # Produto
    prod = cur.execute(
        "SELECT descricao, unidade, categoria FROM produtos WHERE id_produto=?", (id_produto,)
    ).fetchone()
    if not prod:
        conn.close()
        raise HTTPException(404, "Produto nao encontrado")
    # Saldo atual
    saldo = cur.execute(
        "SELECT COALESCE(saldo,0) as s FROM saldos WHERE id_produto=?", (id_produto,)
    ).fetchone()
    # Saidas por mes (ultimos 6 meses)
    saidas_mes = cur.execute("""
        SELECT strftime('%Y-%m', data) as mes,
               SUM(quantidade) as total
        FROM movimentacoes
        WHERE id_produto=? AND tipo='saida'
          AND data >= date('now','-6 months')
        GROUP BY mes
        ORDER BY mes
    """, (id_produto,)).fetchall()
    # Entradas por mes
    entradas_mes = cur.execute("""
        SELECT strftime('%Y-%m', data) as mes,
               SUM(quantidade) as total
        FROM movimentacoes
        WHERE id_produto=? AND tipo IN ('entrada','ajuste')
          AND data >= date('now','-6 months')
        GROUP BY mes
        ORDER BY mes
    """, (id_produto,)).fetchall()
    # Todas movimentacoes recentes
    movs = cur.execute("""
        SELECT tipo, quantidade, responsavel, obs, data
        FROM movimentacoes
        WHERE id_produto=?
        ORDER BY id DESC LIMIT 30
    """, (id_produto,)).fetchall()
    conn.close()
    return {
        "id_produto":   id_produto,
        "descricao":    prod["descricao"],
        "unidade":      prod["unidade"] or "un",
        "categoria":    prod["categoria"] or "GERAL",
        "saldo_atual":  round(float(saldo["s"]) if saldo else 0, 2),
        "saidas_mes":   [{"mes": r["mes"], "total": round(float(r["total"]),2)} for r in saidas_mes],
        "entradas_mes": [{"mes": r["mes"], "total": round(float(r["total"]),2)} for r in entradas_mes],
        "movimentacoes": [dict(r) for r in movs],
    }


@router.get("/relatorio-consumo")
def relatorio_consumo(de: str = "", ate: str = "", categoria: str = "", authorization: str = Header("")):
    verificar_token(authorization)
    conn = get_db()
    cur  = conn.cursor()
    cur.execute("""
        SELECT p.id_produto, p.descricao, p.categoria, p.unidade,
               SUM(m.quantidade) as total_saida,
               COUNT(m.id) as qtd_movs
        FROM movimentacoes m
        JOIN produtos p ON p.id_produto = m.id_produto
        WHERE m.tipo = 'saida'
          AND (? = '' OR m.data >= ?)
          AND (? = '' OR m.data <= ?)
          AND (? = '' OR p.categoria = ?)
        GROUP BY p.id_produto, p.descricao, p.categoria, p.unidade
        HAVING total_saida > 0
        ORDER BY total_saida DESC
    """, (de, de, ate, ate, categoria, categoria))
    rows = cur.fetchall()
    conn.close()
    return {
        "itens": [{
            "id_produto":   r["id_produto"],
            "descricao":    r["descricao"],
            "categoria":    r["categoria"] or "GERAL",
            "unidade":      r["unidade"] or "un",
            "total_saida":  round(float(r["total_saida"]), 2),
            "qtd_movs":     r["qtd_movs"],
        } for r in rows],
        "total": len(rows)
    }

# ── EXPORT EXCEL ─────────────────────────────────────────────────────────────
def _gerar_excel(itens, nome: str):
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    import io, datetime

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = nome[:31]

    # Estilos
    header_fill = PatternFill("solid", fgColor="1C2330")
    header_font = Font(bold=True, color="00D4FF", size=10)
    alt_fill = PatternFill("solid", fgColor="161B22")
    normal_fill = PatternFill("solid", fgColor="0F1117")
    border = Border(bottom=Side(style='thin', color="2D3748"))

    # Cabeçalho
    headers = ["ID", "Descrição", "Categoria", "Unidade", "Saldo", "Saída 90d", "Cons./Dia", "Cobertura (dias)", "Est.Min", "Status"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Larguras
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 12

    status_colors = {"CRÍTICO": "FF4D6A", "ALERTA": "FFB83F", "NORMAL": "00E5A0", "SEM MOVIMENTO": "64748B"}

    for i, item in enumerate(itens):
        saldo = float(item.get("saldo", 0))
        saida = float(item.get("saida_periodo", 0))
        cons_dia = round(saida / 90, 2) if saida else 0
        cobertura = round(saldo / cons_dia) if cons_dia > 0 else 999
        est_min = float(item.get("estoque_minimo", 0))

        if saldo <= 0: status = "SEM MOVIMENTO"
        elif cons_dia > 0 and cobertura <= 15: status = "CRÍTICO"
        elif cons_dia > 0 and cobertura <= 30: status = "ALERTA"
        else: status = "NORMAL"

        row = [
            item.get("id_produto", ""),
            item.get("descricao", ""),
            item.get("categoria", ""),
            item.get("unidade", ""),
            saldo,
            saida,
            cons_dia,
            cobertura if cobertura < 999 else "∞",
            est_min,
            status
        ]
        ws.append(row)

        fill = alt_fill if i % 2 == 0 else normal_fill
        for cell in ws[i+2]:
            cell.fill = fill
            cell.font = Font(color="E2E8F0", size=10)
            cell.border = border

        # Colorir status
        status_cell = ws.cell(row=i+2, column=10)
        cor = status_colors.get(status, "E2E8F0")
        status_cell.font = Font(color=cor, bold=True, size=10)

    # Rodapé
    ws.append([])
    ws.append([f"Gerado em: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nome}.xlsx"}
    )

@router.get("/casa/excel")
def excel_casa(de: str = "", ate: str = "", authorization: str = Header("")):
    verificar_token(authorization)
    return _gerar_excel(get_itens("%CASA%", de, ate), "estoque_casa")

@router.get("/infra/excel")
def excel_infra(de: str = "", ate: str = "", authorization: str = Header("")):
    verificar_token(authorization)
    return _gerar_excel(get_itens("%INFRA%", de, ate), "estoque_infra")

@router.get("/todos/excel")
def excel_todos(de: str = "", ate: str = "", authorization: str = Header("")):
    verificar_token(authorization)
    return _gerar_excel(get_itens("%%", de, ate), "estoque_completo")
